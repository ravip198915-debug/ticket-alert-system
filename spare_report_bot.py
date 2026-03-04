import telebot
import pandas as pd
import os
import schedule
import time
import threading

from datetime import datetime
from telebot.types import ReplyKeyboardMarkup, KeyboardButton
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BOT_TOKEN = "8730404668:AAH4DByLGuRbfJ8gVHjPpKINwl5WPGvRqaA"
MANAGER_ID = 123456789   # manager telegram id

bot = telebot.TeleBot(BOT_TOKEN)

user = {}

dropdown_file = "dropdown_details.xlsx"
report_file = "Spare_Report.xlsx"

# ---------------- LOAD DROPDOWN DATA ---------------- #

data = pd.read_excel(dropdown_file)

data["District"] = data["District"].astype(str).str.strip()
data["Taluk"] = data["Taluk"].astype(str).str.strip()
data["FPS Code"] = data["FPS Code"].astype(str).str.strip()

districts = sorted(data["District"].dropna().unique())
spares = sorted(data["Name of Spare Replaced"].dropna().unique())

# ---------------- CREATE REPORT FILE ---------------- #

if not os.path.exists(report_file):

    columns = [
    "S.No","District","Taluk","FPS Code","Ticket Number",
    "Name of Spare Replaced","Model","Old Serial Number","New Serial Number",
    "Date of Replaced","Name of TE","Remarks",
    "POS Device","Charger","Battery","Camera","Touch",
    "Scanning Glass","Biometric","Biometric Cable"
    ]

    df = pd.DataFrame(columns=columns)
    df.to_excel(report_file,index=False)

# ---------------- START ---------------- #

@bot.message_handler(commands=['start','report'])
def start(message):

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    for d in districts:
        markup.add(KeyboardButton(d))

    user[message.chat.id] = {"step":"district","data":[]}

    bot.send_message(message.chat.id,"Select District",reply_markup=markup)

# ---------------- MAIN FLOW ---------------- #

@bot.message_handler(func=lambda m: True)
def handler(message):

    chat = message.chat.id
    text = message.text.strip()

    if chat not in user:
        return

    step = user[chat]["step"]

    # DISTRICT
    if step == "district":

        user[chat]["data"].append(text)

        taluks = data[data["District"] == text]["Taluk"].unique()

        markup = ReplyKeyboardMarkup(resize_keyboard=True)

        for t in taluks:
            markup.add(KeyboardButton(str(t)))

        user[chat]["step"] = "taluk"

        bot.send_message(chat,"Select Taluk",reply_markup=markup)

    # TALUK
    elif step == "taluk":

        user[chat]["data"].append(text)

        fps = data[data["Taluk"] == text]["FPS Code"].unique()

        markup = ReplyKeyboardMarkup(resize_keyboard=True)

        for f in fps:
            markup.add(KeyboardButton(str(f)))

        user[chat]["step"] = "fps"

        bot.send_message(chat,"Select FPS Code",reply_markup=markup)

    # FPS
    elif step == "fps":

        user[chat]["data"].append(text)

        bot.send_message(chat,"Enter Ticket Number")

        user[chat]["step"] = "ticket"

    # TICKET
    elif step == "ticket":

        if ticket_exists(text):

            bot.send_message(chat,"❌ Ticket already exists. Enter another ticket.")
            return

        user[chat]["data"].append(text)

        markup = ReplyKeyboardMarkup(resize_keyboard=True)

        for s in spares:
            markup.add(KeyboardButton(s))

        user[chat]["step"] = "spare"

        bot.send_message(chat,"Select Spare Replaced",reply_markup=markup)

    # SPARE
    elif step == "spare":

        user[chat]["data"].append(text)

        bot.send_message(chat,"Enter Model")

        user[chat]["step"] = "model"

    # MODEL
    elif step == "model":

        user[chat]["data"].append(text)

        bot.send_message(chat,"Enter Old Serial Number")

        user[chat]["step"] = "old_serial"

    # OLD SERIAL
    elif step == "old_serial":

        user[chat]["data"].append(text)

        bot.send_message(chat,"Enter New Serial Number")

        user[chat]["step"] = "new_serial"

    # NEW SERIAL
    elif step == "new_serial":

        user[chat]["data"].append(text)

        bot.send_message(chat,"Enter Name of TE")

        user[chat]["step"] = "te"

    # TE
    elif step == "te":

        user[chat]["data"].append(text)

        bot.send_message(chat,"Enter Remarks")

        user[chat]["step"] = "remarks"

    # REMARKS
    elif step == "remarks":

        user[chat]["data"].append(text)

        save_report(user[chat]["data"])

        bot.send_message(chat,"✅ Spare Replacement Report Saved")

        del user[chat]

# ---------------- DUPLICATE CHECK ---------------- #

def ticket_exists(ticket):

    if not os.path.exists(report_file):
        return False

    df = pd.read_excel(report_file)

    return ticket in df["Ticket Number"].astype(str).values

# ---------------- SAVE REPORT ---------------- #

def save_report(data):

    df = pd.read_excel(report_file)

    sno = len(df) + 1

    row = {
    "S.No":sno,
    "District":data[0],
    "Taluk":data[1],
    "FPS Code":data[2],
    "Ticket Number":data[3],
    "Name of Spare Replaced":data[4],
    "Model":data[5],
    "Old Serial Number":data[6],
    "New Serial Number":data[7],
    "Date of Replaced":datetime.today().strftime("%d-%b-%Y"),
    "Name of TE":data[8],
    "Remarks":data[9],
    "POS Device":"",
    "Charger":"",
    "Battery":"",
    "Camera":"",
    "Touch":"",
    "Scanning Glass":"",
    "Biometric":"",
    "Biometric Cable":""
    }

    spare_map = {
        "Device":"POS Device",
        "Charger":"Charger",
        "Battery":"Battery",
        "Camera":"Camera",
        "Touch":"Touch",
        "Scanning Glass":"Scanning Glass",
        "Biometric":"Biometric",
        "Biometric Cable":"Biometric Cable"
    }

    if data[4] in spare_map:
        row[spare_map[data[4]]] = 1

    df = df._append(row,ignore_index=True)

    try:
        df.to_excel(report_file,index=False)
        format_excel()
    except PermissionError:
        print("Excel file is open. Please close it.")

# ---------------- FORMAT EXCEL ---------------- #

def format_excel():

    wb = load_workbook(report_file)
    ws = wb.active

    fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    wb.save(report_file)

# ---------------- DAILY REPORT ---------------- #

def send_daily_report():

    df = pd.read_excel(report_file)

    today = datetime.today().strftime("%d-%b-%Y")

    df_today = df[df["Date of Replaced"] == today]

    if df_today.empty:
        return

    summary = df_today.groupby("District").size()

    msg = "📊 Daily Spare Replacement Report\n\n"

    for d,c in summary.items():
        msg += f"{d} : {c}\n"

    bot.send_message(MANAGER_ID,msg)

# ---------------- SCHEDULER ---------------- #

def scheduler():

    schedule.every().day.at("18:00").do(send_daily_report)

    while True:
        schedule.run_pending()
        time.sleep(30)

threading.Thread(target=scheduler).start()

# ---------------- RUN BOT ---------------- #

bot.infinity_polling()